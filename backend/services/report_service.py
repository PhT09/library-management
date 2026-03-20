from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date
from typing import Optional
import io
import csv
import models

def get_top_borrowed_books(db: Session, limit: int = 10, start_date: Optional[date] = None, end_date: Optional[date] = None):
    """Đếm số lượng mượn đối với mỗi đầu sách, sắp xếp giảm dần"""
    query = (
        db.query(models.Book, func.count(models.BorrowRecord.id).label("borrow_count"))
        .join(models.BookCopy, models.Book.id == models.BookCopy.book_id)
        .join(models.BorrowRecord, models.BookCopy.id == models.BorrowRecord.book_copy_id)
    )
    
    if start_date:
        query = query.filter(models.BorrowRecord.borrow_date >= start_date)
    if end_date:
        query = query.filter(models.BorrowRecord.borrow_date <= end_date)
        
    query = (
        query.group_by(models.Book.id)
        .order_by(func.count(models.BorrowRecord.id).desc())
        .limit(limit)
    )
    
    results = query.all()
    
    report = []
    for book, count in results:
        category = db.query(models.Category).filter(models.Category.id == book.category_id).first()
        report.append({
            "book_id": book.id,
            "book_name": book.name,
            "author": book.author,
            "category_id": book.category_id,
            "category_name": category.name if category else None,
            "borrow_count": count
        })
        
    return report

def get_unreturned_readers(db: Session, start_date: Optional[date] = None, end_date: Optional[date] = None):
    """Danh sách những sinh viên có phiếu mượn đang ở trạng thái Chưa Trả"""
    query = (
        db.query(models.BorrowRecord, models.Reader, models.BookCopy, models.Book)
        .join(models.Reader, models.BorrowRecord.reader_id == models.Reader.id)
        .join(models.BookCopy, models.BorrowRecord.book_copy_id == models.BookCopy.id)
        .join(models.Book, models.BookCopy.book_id == models.Book.id)
        .filter(models.BorrowRecord.status == "Borrowing")
    )
    
    if start_date:
        query = query.filter(models.BorrowRecord.borrow_date >= start_date)
    if end_date:
        query = query.filter(models.BorrowRecord.borrow_date <= end_date)
    
    results = query.all()
    
    report = []
    for record, reader, copy, book in results:
        report.append({
            "borrow_id": record.id,
            "reader_id": reader.id,
            "reader_name": reader.full_name,
            "class_name": reader.class_name,
            "book_copy_id": copy.id,
            "book_name": book.name,
            "borrow_date": record.borrow_date.strftime("%Y-%m-%d") if record.borrow_date else None
        })
        
    return report

# ====================== XUẤT FILE BÁO CÁO ======================

def export_to_csv(data: list, report_type: str) -> io.StringIO:
    """Xuất báo cáo ra file CSV"""
    output = io.StringIO()
    
    if report_type == "top_books":
        headers = ["STT", "Mã sách", "Tên sách", "Tác giả", "Chuyên ngành", "Số lượt mượn"]
        writer = csv.writer(output)
        writer.writerow(headers)
        for idx, item in enumerate(data, 1):
            writer.writerow([
                idx,
                item.get("book_id", ""),
                item.get("book_name", ""),
                item.get("author", ""),
                item.get("category_name", ""),
                item.get("borrow_count", 0)
            ])
    elif report_type == "unreturned_readers":
        headers = ["STT", "Mã phiếu", "Mã ĐG", "Họ tên", "Lớp", "Mã bản sao", "Tên sách", "Ngày mượn"]
        writer = csv.writer(output)
        writer.writerow(headers)
        for idx, item in enumerate(data, 1):
            writer.writerow([
                idx,
                item.get("borrow_id", ""),
                item.get("reader_id", ""),
                item.get("reader_name", ""),
                item.get("class_name", ""),
                item.get("book_copy_id", ""),
                item.get("book_name", ""),
                item.get("borrow_date", "")
            ])
    elif report_type == "readers":
        headers = ["STT", "Mã ĐG", "Họ Tên", "Lớp", "Ngày sinh", "Giới tính", "Trạng thái"]
        writer = csv.writer(output)
        writer.writerow(headers)
        for idx, item in enumerate(data, 1):
            writer.writerow([
                idx,
                item.get("id", ""),
                item.get("full_name", ""),
                item.get("class_name", ""),
                item.get("dob", ""),
                item.get("gender", ""),
                item.get("is_active", "")
            ])
    
    output.seek(0)
    return output

def export_to_excel(data: list, report_type: str) -> io.BytesIO:
    """Xuất báo cáo ra file Excel (.xlsx)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise Exception("Thư viện openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if report_type == "top_books":
        ws.title = "Top Sách Mượn Nhiều"
        headers = ["STT", "Mã sách", "Tên sách", "Tác giả", "Chuyên ngành", "Số lượt mượn"]
        
        # Tiêu đề báo cáo
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "BÁO CÁO TOP SÁCH ĐƯỢC MƯỢN NHIỀU NHẤT"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dữ liệu
        for idx, item in enumerate(data, 1):
            row = idx + 3
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=item.get("book_id", "")).border = thin_border
            ws.cell(row=row, column=3, value=item.get("book_name", "")).border = thin_border
            ws.cell(row=row, column=4, value=item.get("author", "")).border = thin_border
            ws.cell(row=row, column=5, value=item.get("category_name", "")).border = thin_border
            ws.cell(row=row, column=6, value=item.get("borrow_count", 0)).border = thin_border
        
        # Auto-fit column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
            
    elif report_type == "unreturned_readers":
        ws.title = "ĐG Chưa Trả Sách"
        headers = ["STT", "Mã phiếu", "Mã ĐG", "Họ tên", "Lớp", "Mã bản sao", "Tên sách", "Ngày mượn"]
        
        # Tiêu đề báo cáo
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "BÁO CÁO DANH SÁCH ĐỌC GIẢ CHƯA TRẢ SÁCH"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dữ liệu
        for idx, item in enumerate(data, 1):
            row = idx + 3
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=item.get("borrow_id", "")).border = thin_border
            ws.cell(row=row, column=3, value=item.get("reader_id", "")).border = thin_border
            ws.cell(row=row, column=4, value=item.get("reader_name", "")).border = thin_border
            ws.cell(row=row, column=5, value=item.get("class_name", "")).border = thin_border
            ws.cell(row=row, column=6, value=item.get("book_copy_id", "")).border = thin_border
            ws.cell(row=row, column=7, value=item.get("book_name", "")).border = thin_border
            ws.cell(row=row, column=8, value=item.get("borrow_date", "")).border = thin_border
        
        # Auto-fit column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 14
        
    elif report_type == "readers":
        ws.title = "Danh sách Độc giả"
        headers = ["STT", "Mã ĐG", "Họ Tên", "Lớp", "Ngày sinh", "Giới tính", "Trạng thái"]
        
        # Tiêu đề báo cáo
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "DANH SÁCH ĐỘC GIẢ THƯ VIỆN"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dữ liệu
        for idx, item in enumerate(data, 1):
            row = idx + 3
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=item.get("id", "")).border = thin_border
            ws.cell(row=row, column=3, value=item.get("full_name", "")).border = thin_border
            ws.cell(row=row, column=4, value=item.get("class_name", "")).border = thin_border
            ws.cell(row=row, column=5, value=item.get("dob", "")).border = thin_border
            ws.cell(row=row, column=6, value=item.get("gender", "")).border = thin_border
            status_text = "Hoạt động" if item.get("is_active") else "Đã khóa"
            ws.cell(row=row, column=7, value=status_text).border = thin_border
        
        # Auto-fit column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
